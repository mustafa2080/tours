from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from modeltranslation.admin import TranslationAdmin, TranslationTabularInline
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
import pandas as pd
import csv
import io
import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.text import slugify
from django.utils.safestring import mark_safe # Import mark_safe at top level
from .forms import TourImportForm, CategoryImportForm # Import CategoryImportForm
from .models import (
    Destination, DestinationImage, Category, Activity, Tour, TourImage,
    TourDate, TourGuide, TourItinerary, TourFAQ, Promotion
)

# Ensure translation registration happens before admin classes are defined
import tour.translation

# Admin classes follow
class DestinationImageInline(admin.TabularInline):
    model = DestinationImage
    extra = 1


@admin.register(Destination)
class DestinationAdmin(TranslationAdmin):
    list_display = ('name', 'country', 'city', 'is_featured', 'is_active')
    list_filter = ('country', 'is_featured', 'is_active')
    search_fields = ('name', 'country', 'city')
    prepopulated_fields = {'slug': ('name',)}
    inlines = [DestinationImageInline]


@admin.register(Category)
class CategoryAdmin(TranslationAdmin):
    list_display = ('name', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('name',)
    prepopulated_fields = {'slug': ('name',)}

    # Add import button to changelist view
    change_list_template = 'admin/tour/category/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-categories/', self.import_categories, name='import_categories'),
            path('download-sample-category-file/', self.download_sample_file, name='download_sample_category_file'),
        ]
        return my_urls + urls

    def download_sample_file(self, request):
        """Generate and download a sample Excel file for category import"""
        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create Categories sheet
        categories_sheet = wb.active
        categories_sheet.title = "Categories"

        # Add headers for Categories sheet
        category_headers = [
            'name_en', 'name_ar', 'name_fr', 'name_de',
            'slug',
            'description_en', 'description_ar', 'description_fr', 'description_de',
            'is_active'
        ]
        categories_sheet.append(category_headers)

        # Add sample data for Categories
        categories_sheet.append([
            'Adventure Tours', 'جولات المغامرة', 'Circuits Aventure', 'Abenteuertouren',
            'adventure-tours',
            'Exciting adventure tours for thrill seekers.',
            'جولات مغامرة مثيرة لعشاق الإثارة.',
            'Circuits d\'aventure passionnants pour les amateurs de sensations fortes.',
            'Aufregende Abenteuertouren für Adrenalin-Junkies.',
            'True'
        ])

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sample_categories_import.xlsx"'

        # Save the workbook to the response
        wb.save(response)
        return response

    def _process_categories_data(self, data, update_existing, categories_created, categories_updated, errors):
        """Process category data from the imported file"""
        for index, row in data.iterrows():
            try:
                # Define actual CSV header names from categories_data.csv
                csv_name_en_col = 'Category Name [en]*'
                csv_name_ar_col = 'Category Name [ar]'
                csv_name_fr_col = 'Category Name [fr]'
                csv_name_de_col = 'Category Name [de]'
                csv_slug_col = 'Slug'
                csv_desc_en_col = 'Description [en]'
                csv_desc_ar_col = 'Description [ar]'
                csv_desc_fr_col = 'Description [fr]'
                csv_desc_de_col = 'Description [de]'

                # Check required fields using the correct header from CSV
                if csv_name_en_col not in row or pd.isna(row[csv_name_en_col]):
                    errors.append(f"Row {index+1}: Missing required field '{csv_name_en_col}'")
                    continue

                name_en = row[csv_name_en_col] # Store the English name

                # Generate slug if not provided in CSV
                slug = row.get(csv_slug_col, '')
                if not slug:
                    slug = slugify(name_en)

                # Check if category exists (for update) using the English name
                category = None
                if update_existing:
                    try:
                        category = Category.objects.get(name=name_en) # Find by English name
                    except Category.DoesNotExist:
                        pass

                # Create or update category
                if category:
                    # Update existing category
                    category.slug = slug

                    # Update translatable fields using correct CSV headers
                    if csv_name_en_col in row and not pd.isna(row[csv_name_en_col]):
                        category.name_en = row[csv_name_en_col] # Already have name_en
                    if csv_name_ar_col in row and not pd.isna(row[csv_name_ar_col]):
                        category.name_ar = row[csv_name_ar_col]
                    if csv_name_fr_col in row and not pd.isna(row[csv_name_fr_col]):
                        category.name_fr = row[csv_name_fr_col]
                    if csv_name_de_col in row and not pd.isna(row[csv_name_de_col]):
                        category.name_de = row[csv_name_de_col]

                    if csv_desc_en_col in row and not pd.isna(row[csv_desc_en_col]):
                        category.description_en = row[csv_desc_en_col]
                    if csv_desc_ar_col in row and not pd.isna(row[csv_desc_ar_col]):
                        category.description_ar = row[csv_desc_ar_col]
                    if csv_desc_fr_col in row and not pd.isna(row[csv_desc_fr_col]):
                        category.description_fr = row[csv_desc_fr_col]
                    if csv_desc_de_col in row and not pd.isna(row[csv_desc_de_col]):
                        category.description_de = row[csv_desc_de_col]

                    # Update other fields (is_active is not in categories_data.csv, keep existing or default)
                    if 'is_active' in row and not pd.isna(row['is_active']):
                        category.is_active = str(row['is_active']).lower() in ('true', 'yes', '1')

                    category.save()
                    categories_updated += 1
                else:
                    # Create new category with required fields using correct headers
                    category_data = {
                        'name': name_en,  # Use name_en variable from CSV
                        'slug': slug,
                        'is_active': True # Default is_active to True for new categories
                    }

                    # Add translatable fields using correct CSV headers
                    if csv_name_ar_col in row and not pd.isna(row[csv_name_ar_col]):
                        category_data['name_ar'] = row[csv_name_ar_col]
                    if csv_name_fr_col in row and not pd.isna(row[csv_name_fr_col]):
                        category_data['name_fr'] = row[csv_name_fr_col]
                    if csv_name_de_col in row and not pd.isna(row[csv_name_de_col]):
                        category_data['name_de'] = row[csv_name_de_col]

                    if csv_desc_en_col in row and not pd.isna(row[csv_desc_en_col]):
                        category_data['description_en'] = row[csv_desc_en_col]
                    if csv_desc_ar_col in row and not pd.isna(row[csv_desc_ar_col]):
                        category_data['description_ar'] = row[csv_desc_ar_col]
                    if csv_desc_fr_col in row and not pd.isna(row[csv_desc_fr_col]):
                        category_data['description_fr'] = row[csv_desc_fr_col]
                    if csv_desc_de_col in row and not pd.isna(row[csv_desc_de_col]):
                        category_data['description_de'] = row[csv_desc_de_col]

                    # Add optional fields (is_active is not in categories_data.csv, defaults to True)
                    if 'is_active' in row and not pd.isna(row['is_active']):
                        category_data['is_active'] = str(row['is_active']).lower() in ('true', 'yes', '1')

                    category = Category(**category_data)
                    category.save()
                    categories_created += 1

            except Exception as e:
                errors.append(f"Row {index+1}: {str(e)}")

        return categories_created, categories_updated

    def import_categories(self, request):
        if request.method == "POST":
            form = CategoryImportForm(request.POST, request.FILES)
            if form.is_valid():
                file_format = form.cleaned_data['file_format']
                category_file = request.FILES['file']
                update_existing = form.cleaned_data['update_existing']

                # Check file extension
                if file_format == 'csv' and not category_file.name.endswith('.csv'):
                    messages.error(request, _("Wrong file format. Please upload a CSV file."))
                    return redirect('admin:tour_category_changelist')

                if file_format == 'xlsx' and not category_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, _("Wrong file format. Please upload an Excel file."))
                    return redirect('admin:tour_category_changelist')

                # Process the file
                try:
                    # Start transaction to ensure all-or-nothing import
                    with transaction.atomic():
                        categories_created = 0
                        categories_updated = 0
                        errors = []

                        # Read the file
                        if file_format == 'csv':
                            # CSV file
                            data = pd.read_csv(category_file, encoding='utf-8')
                        else:
                            # Excel file
                            data = pd.read_excel(category_file)

                        # Process the data
                        categories_created, categories_updated = self._process_categories_data(
                            data,
                            update_existing,
                            categories_created,
                            categories_updated,
                            errors
                        )

                        # Show success message
                        success_messages = []
                        if categories_created > 0:
                            success_messages.append(_(f"{categories_created} categories created"))
                        if categories_updated > 0:
                            success_messages.append(_(f"{categories_updated} categories updated"))

                        if success_messages:
                            # Convert lazy objects to strings before joining (consistency fix)
                            messages.success(request, str(_("Import successful: ")) + ", ".join(map(str, success_messages)))
                        else:
                            messages.warning(request, _("Import completed, but no data was imported."))

                        # Show errors if any
                        if errors:
                            from django.utils.safestring import mark_safe # Re-import locally as a workaround
                            error_message = _("Errors occurred during import:") + "<br>" + "<br>".join(errors)
                            messages.error(request, mark_safe(error_message))

                except Exception as e:
                    messages.error(request, _(f"Error processing file: {str(e)}"))

                return redirect('admin:tour_category_changelist')
        else:
            form = CategoryImportForm()

        context = {
            'form': form,
            'title': _("Import Categories"),
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
        }
        return render(request, 'admin/tour/category/import_form.html', context)


@admin.register(Activity)
class ActivityAdmin(TranslationAdmin):
    list_display = ('name',)
    search_fields = ('name',)


class TourImageInline(admin.TabularInline):
    model = TourImage
    extra = 1


class TourItineraryInline(TranslationTabularInline):
    model = TourItinerary
    extra = 1


class TourFAQInline(TranslationTabularInline):
    model = TourFAQ
    extra = 1


class TourDateInline(admin.TabularInline):
    model = TourDate
    extra = 1
    fields = ('start_date', 'end_date', 'available_seats', 'is_active')


@admin.register(Tour)
class TourAdmin(TranslationAdmin):
    list_display = ('name', 'destination', 'duration_days', 'price', 'is_featured', 'is_active')
    list_filter = ('destination', 'categories', 'is_featured', 'is_active')
    search_fields = ('name', 'destination__name', 'short_description')
    prepopulated_fields = {'slug': ('name',)}
    filter_horizontal = ('categories', 'activities')
    inlines = [TourImageInline, TourItineraryInline, TourFAQInline, TourDateInline]

    # Add import button to changelist view
    change_list_template = 'admin/tour/tour/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-tours/', self.import_tours, name='import_tours'),
            path('download-sample-file/', self.download_sample_file, name='download_sample_file'),
        ]
        return my_urls + urls

    def download_sample_file(self, request):
        """Generate and download a sample Excel file for tour import"""
        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create Tours sheet
        tours_sheet = wb.active
        tours_sheet.title = "Tours"

        # Add headers for Tours sheet
        tour_headers = [
            'name_en', 'name_ar', 'name_fr', 'name_de',
            'slug',
            'description_en', 'description_ar', 'description_fr', 'description_de',
            'short_description_en', 'short_description_ar', 'short_description_fr', 'short_description_de',
            'destination', 'categories', 'price', 'discount_price', 'duration_days', 'duration_nights', # Added duration_nights
            'max_group_size', 'min_age', 'is_featured', 'is_active'
        ]
        tours_sheet.append(tour_headers)

        # Add sample data for Tours
        tours_sheet.append([
            'Cairo Pyramids Tour', 'جولة أهرامات الجيزة', 'Visite des Pyramides du Caire', 'Kairo-Pyramiden-Tour',
            'cairo-pyramids-tour',
            'Explore the ancient wonders of Egypt with our guided tour of the Pyramids.',
            'استكشف عجائب مصر القديمة مع جولتنا المصحوبة بمرشدين في الأهرامات.',
            'Explorez les merveilles anciennes de l\'Égypte avec notre visite guidée des Pyramides.',
            'Entdecken Sie die antiken Wunder Ägyptens mit unserer geführten Tour zu den Pyramiden.',
            'A 3-day tour of Cairo and the Pyramids',
            'جولة لمدة 3 أيام في القاهرة والأهرامات',
            'Une visite de 3 jours au Caire et aux Pyramides',
            'Eine 3-tägige Tour durch Kairo und die Pyramiden',
            'Cairo', 'Historical,Cultural', '1200', '999', '3', '2', # Added sample duration_nights (e.g., 2)
            '15', '8', 'easy', 'True', 'True'
        ])

        # Create Itinerary sheet
        itinerary_sheet = wb.create_sheet(title="Itinerary")

        # Add headers for Itinerary sheet
        itinerary_headers = [
            'tour_name_en', 'day_number',
            'title_en', 'title_ar', 'title_fr', 'title_de',
            'description_en', 'description_ar', 'description_fr', 'description_de',
            'meals_en', 'meals_ar', 'meals_fr', 'meals_de',
            'accommodation_en', 'accommodation_ar', 'accommodation_fr', 'accommodation_de'
        ]
        itinerary_sheet.append(itinerary_headers)

        # Add sample data for Itinerary
        itinerary_sheet.append([
            'Cairo Pyramids Tour', '1',
            'Pyramids of Giza', 'أهرامات الجيزة', 'Pyramides de Gizeh', 'Pyramiden von Gizeh',
            'Visit the Great Pyramids and the Sphinx.',
            'زيارة الأهرامات العظيمة وأبو الهول.',
            'Visitez les Grandes Pyramides et le Sphinx.',
            'Besuchen Sie die Großen Pyramiden und die Sphinx.',
            'Breakfast, Lunch', 'الإفطار والغداء', 'Petit-déjeuner, Déjeuner', 'Frühstück, Mittagessen',
            'Marriott Hotel', 'فندق ماريوت', 'Hôtel Marriott', 'Marriott Hotel'
        ])

        itinerary_sheet.append([
            'Cairo Pyramids Tour', '2',
            'Egyptian Museum', 'المتحف المصري', 'Musée égyptien', 'Ägyptisches Museum',
            'Explore the treasures of ancient Egypt at the Egyptian Museum.',
            'استكشف كنوز مصر القديمة في المتحف المصري.',
            'Explorez les trésors de l\'Égypte ancienne au Musée égyptien.',
            'Entdecken Sie die Schätze des alten Ägypten im Ägyptischen Museum.',
            'Breakfast, Lunch, Dinner', 'الإفطار والغداء والعشاء', 'Petit-déjeuner, Déjeuner, Dîner', 'Frühstück, Mittagessen, Abendessen',
            'Marriott Hotel', 'فندق ماريوت', 'Hôtel Marriott', 'Marriott Hotel'
        ])

        # Create FAQs sheet
        faqs_sheet = wb.create_sheet(title="FAQs")

        # Add headers for FAQs sheet
        faq_headers = [
            'tour_name_en',
            'question_en', 'question_ar', 'question_fr', 'question_de',
            'answer_en', 'answer_ar', 'answer_fr', 'answer_de',
            'display_order'
        ]
        faqs_sheet.append(faq_headers)

        # Add sample data for FAQs
        faqs_sheet.append([
            'Cairo Pyramids Tour',
            'What should I wear?', 'ماذا يجب أن ألبس؟', 'Que dois-je porter?', 'Was soll ich anziehen?',
            'Comfortable clothing and walking shoes are recommended. Also bring a hat and sunscreen.',
            'يوصى بارتداء ملابس مريحة وأحذية مشي. أحضر أيضًا قبعة وواقي من الشمس.',
            'Des vêtements confortables et des chaussures de marche sont recommandés. Apportez également un chapeau et de la crème solaire.',
            'Bequeme Kleidung und Wanderschuhe werden empfohlen. Bringen Sie auch einen Hut und Sonnencreme mit.',
            '1'
        ])

        faqs_sheet.append([
            'Cairo Pyramids Tour',
            'Is photography allowed?', 'هل التصوير مسموح به؟', 'La photographie est-elle autorisée?', 'Ist Fotografieren erlaubt?',
            'Photography is allowed in most areas, but some museums may charge an additional fee for cameras.',
            'التصوير مسموح به في معظم المناطق، ولكن بعض المتاحف قد تفرض رسومًا إضافية للكاميرات.',
            'La photographie est autorisée dans la plupart des zones, mais certains musées peuvent facturer des frais supplémentaires pour les caméras.',
            'Fotografieren ist in den meisten Bereichen erlaubt, aber einige Museen erheben möglicherweise eine zusätzliche Gebühr für Kameras.',
            '2'
        ])

        # Create a response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=tour_import_sample.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response

    def _process_tours_data(self, data, create_missing_destinations, create_missing_categories,
                           update_existing, tours_created, tours_updated, errors):
        """Process tour data from the imported file"""
        for index, row in data.iterrows():
            try:
                # Check required fields
                if 'name_en' not in row or pd.isna(row['name_en']):
                    errors.append(f"Row {index+1}: Missing required field 'name_en'")
                    continue

                if 'destination' not in row or pd.isna(row['destination']):
                    errors.append(f"Row {index+1}: Missing required field 'destination'")
                    continue

                # Get or create destination
                destination_name = row['destination']
                try:
                    destination = Destination.objects.get(name=destination_name)
                except Destination.DoesNotExist:
                    if create_missing_destinations:
                        # Create new destination
                        destination = Destination(
                            name=destination_name,
                            slug=slugify(destination_name),
                            country=row.get('country', ''),
                            city=row.get('city', ''),
                            is_active=True
                        )
                        destination.save()
                    else:
                        errors.append(f"Row {index+1}: Destination '{destination_name}' does not exist")
                        continue

                # Generate initial base slug
                base_slug = row.get('slug', '')
                if not base_slug:
                    base_slug = slugify(row['name_en'])

                # Check if tour exists (for update)
                tour = None
                if update_existing:
                    try:
                        tour = Tour.objects.get(name=row['name_en'])
                    except Tour.DoesNotExist:
                        pass

                # Create or update tour
                if tour:
                    # --- Update existing tour ---

                    # Ensure slug uniqueness before saving
                    final_slug = base_slug
                    counter = 1
                    while Tour.objects.filter(slug=final_slug).exclude(pk=tour.pk).exists():
                        counter += 1
                        final_slug = f"{base_slug}-{counter}"
                    tour.slug = final_slug # Assign unique slug

                    tour.destination = destination

                    # Update translatable fields
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'name_{lang}' in row and not pd.isna(row[f'name_{lang}']):
                            setattr(tour, f'name_{lang}', row[f'name_{lang}'])
                        if f'description_{lang}' in row and not pd.isna(row[f'description_{lang}']):
                            setattr(tour, f'description_{lang}', row[f'description_{lang}'])
                        if f'short_description_{lang}' in row and not pd.isna(row[f'short_description_{lang}']):
                            setattr(tour, f'short_description_{lang}', row[f'short_description_{lang}'])

                    # Update other fields
                    if 'price' in row and not pd.isna(row['price']):
                        tour.price = row['price']
                    if 'discount_price' in row and not pd.isna(row['discount_price']):
                        tour.discount_price = row['discount_price']
                    if 'duration_days' in row and not pd.isna(row['duration_days']):
                        tour.duration_days = row['duration_days']
                    if 'duration_nights' in row and not pd.isna(row['duration_nights']): # Added duration_nights handling
                        tour.duration_nights = row['duration_nights']
                    if 'max_group_size' in row and not pd.isna(row['max_group_size']):
                        tour.max_people = row['max_group_size'] # Changed to max_people
                    if 'min_age' in row and not pd.isna(row['min_age']):
                        tour.min_age = row['min_age']
                    if 'difficulty' in row and not pd.isna(row['difficulty']):
                        tour.difficulty = row['difficulty']
                    if 'is_featured' in row and not pd.isna(row['is_featured']):
                        tour.is_featured = str(row['is_featured']).lower() in ('true', 'yes', '1')
                    if 'is_active' in row and not pd.isna(row['is_active']):
                        tour.is_active = str(row['is_active']).lower() in ('true', 'yes', '1')

                    tour.save()
                    tours_updated += 1
                else:
                    # --- Create new tour ---

                    # Ensure slug uniqueness before creating
                    final_slug = base_slug
                    counter = 1
                    while Tour.objects.filter(slug=final_slug).exists():
                        counter += 1
                        final_slug = f"{base_slug}-{counter}"

                    # Handle max_people robustly before creating tour_data
                    max_people_value = 10 # Default value
                    if 'max_group_size' in row and not pd.isna(row['max_group_size']):
                        try:
                            max_people_value = int(row['max_group_size'])
                        except (ValueError, TypeError):
                            # Log error or use default if conversion fails
                            errors.append(f"Row {index+1}: Invalid value '{row['max_group_size']}' for 'max_group_size'. Using default {max_people_value}.")
                            # Keep the default value

                    # Create new tour with required fields
                    tour_data = {
                        'name': row['name_en'],  # Default name
                        'slug': final_slug, # Assign unique slug
                        'destination': destination,
                        'price': row.get('price', 0),
                        'duration_days': row.get('duration_days', 1),
                        'duration_nights': row.get('duration_nights', 0),
                        'is_active': True,
                        'max_people': max_people_value # Use the processed value
                    }

                    # Add translatable fields
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'name_{lang}' in row and not pd.isna(row[f'name_{lang}']):
                            tour_data[f'name_{lang}'] = row[f'name_{lang}']
                        if f'description_{lang}' in row and not pd.isna(row[f'description_{lang}']):
                            tour_data[f'description_{lang}'] = row[f'description_{lang}']
                        if f'short_description_{lang}' in row and not pd.isna(row[f'short_description_{lang}']):
                            tour_data[f'short_description_{lang}'] = row[f'short_description_{lang}']

                    # Add optional fields (max_people already set in tour_data)
                    if 'discount_price' in row and not pd.isna(row['discount_price']):
                        tour_data['discount_price'] = row['discount_price']
                    # duration_nights already added above
                    # max_people handled during initialization above
                    if 'min_age' in row and not pd.isna(row['min_age']):
                        tour_data['min_age'] = row['min_age']
                    if 'difficulty' in row and not pd.isna(row['difficulty']):
                        tour_data['difficulty'] = row['difficulty']
                    if 'is_featured' in row and not pd.isna(row['is_featured']):
                        tour_data['is_featured'] = str(row['is_featured']).lower() in ('true', 'yes', '1')
                    if 'is_active' in row and not pd.isna(row['is_active']):
                        tour_data['is_active'] = str(row['is_active']).lower() in ('true', 'yes', '1')

                    tour = Tour(**tour_data)
                    tour.save()
                    tours_created += 1

                # Process categories
                if 'categories' in row and not pd.isna(row['categories']):
                    categories = [c.strip() for c in str(row['categories']).split(',')]
                    for cat_name in categories:
                        if cat_name:
                            try:
                                category = Category.objects.get(name=cat_name)
                            except Category.DoesNotExist:
                                if create_missing_categories:
                                    category = Category(
                                        name=cat_name,
                                        slug=slugify(cat_name),
                                        is_active=True
                                    )
                                    category.save()
                                else:
                                    errors.append(f"Row {index+1}: Category '{cat_name}' does not exist")

                            tour.categories.add(category)

                # Process activities
                if 'activities' in row and not pd.isna(row['activities']):
                    activities = [a.strip() for a in str(row['activities']).split(',')]
                    for act_name in activities:
                        if act_name:
                            activity, created = Activity.objects.get_or_create(name=act_name)
                            tour.activities.add(activity)

            except Exception as e:
                errors.append(f"Row {index+1}: {str(e)}")

        return tours_created, tours_updated

    def _process_itinerary_data(self, data, errors):
        """Process itinerary data from the imported file"""
        itineraries_created = 0

        for index, row in data.iterrows():
            try:
                # Check required fields
                if 'tour_name_en' not in row or pd.isna(row['tour_name_en']):
                    errors.append(f"Itinerary Row {index+1}: Missing required field 'tour_name_en'")
                    continue

                if 'day_number' not in row or pd.isna(row['day_number']):
                    errors.append(f"Itinerary Row {index+1}: Missing required field 'day_number'")
                    continue

                if 'title_en' not in row or pd.isna(row['title_en']):
                    errors.append(f"Itinerary Row {index+1}: Missing required field 'title_en'")
                    continue

                # Find the tour
                try:
                    tour = Tour.objects.get(name_en=row['tour_name_en'])
                except Tour.DoesNotExist:
                    errors.append(f"Itinerary Row {index+1}: Tour '{row['tour_name_en']}' does not exist")
                    continue

                # Check if itinerary day exists
                try:
                    itinerary = TourItinerary.objects.get(tour=tour, day=row['day_number'])
                    # Update existing itinerary
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'title_{lang}' in row and not pd.isna(row[f'title_{lang}']):
                            setattr(itinerary, f'title_{lang}', row[f'title_{lang}'])
                        if f'description_{lang}' in row and not pd.isna(row[f'description_{lang}']):
                            setattr(itinerary, f'description_{lang}', row[f'description_{lang}'])
                        if f'meals_{lang}' in row and not pd.isna(row[f'meals_{lang}']):
                            setattr(itinerary, f'meals_{lang}', row[f'meals_{lang}'])
                        if f'accommodation_{lang}' in row and not pd.isna(row[f'accommodation_{lang}']):
                            setattr(itinerary, f'accommodation_{lang}', row[f'accommodation_{lang}'])

                    itinerary.save()
                except TourItinerary.DoesNotExist:
                    # Create new itinerary
                    itinerary_data = {
                        'tour': tour,
                        'day': row['day_number'],
                        'title': row['title_en'],  # Default title
                    }

                    # Add translatable fields
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'title_{lang}' in row and not pd.isna(row[f'title_{lang}']):
                            itinerary_data[f'title_{lang}'] = row[f'title_{lang}']
                        if f'description_{lang}' in row and not pd.isna(row[f'description_{lang}']):
                            itinerary_data[f'description_{lang}'] = row[f'description_{lang}']
                        if f'meals_{lang}' in row and not pd.isna(row[f'meals_{lang}']):
                            itinerary_data[f'meals_{lang}'] = row[f'meals_{lang}']
                        if f'accommodation_{lang}' in row and not pd.isna(row[f'accommodation_{lang}']):
                            itinerary_data[f'accommodation_{lang}'] = row[f'accommodation_{lang}']

                    itinerary = TourItinerary(**itinerary_data)
                    itinerary.save()
                    itineraries_created += 1

            except Exception as e:
                errors.append(f"Itinerary Row {index+1}: {str(e)}")

        return itineraries_created

    def _process_faq_data(self, data, errors):
        """Process FAQ data from the imported file"""
        faqs_created = 0

        for index, row in data.iterrows():
            try:
                # Check required fields
                if 'tour_name_en' not in row or pd.isna(row['tour_name_en']):
                    errors.append(f"FAQ Row {index+1}: Missing required field 'tour_name_en'")
                    continue

                if 'question_en' not in row or pd.isna(row['question_en']):
                    errors.append(f"FAQ Row {index+1}: Missing required field 'question_en'")
                    continue

                if 'answer_en' not in row or pd.isna(row['answer_en']):
                    errors.append(f"FAQ Row {index+1}: Missing required field 'answer_en'")
                    continue

                # Find the tour
                try:
                    tour = Tour.objects.get(name_en=row['tour_name_en'])
                except Tour.DoesNotExist:
                    errors.append(f"FAQ Row {index+1}: Tour '{row['tour_name_en']}' does not exist")
                    continue

                # Get display order
                display_order = 0
                if 'display_order' in row and not pd.isna(row['display_order']):
                    display_order = int(row['display_order'])

                # Check if FAQ exists (by question)
                existing_faq = TourFAQ.objects.filter(tour=tour, question_en=row['question_en']).first()

                if existing_faq:
                    # Update existing FAQ
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'question_{lang}' in row and not pd.isna(row[f'question_{lang}']):
                            setattr(existing_faq, f'question_{lang}', row[f'question_{lang}'])
                        if f'answer_{lang}' in row and not pd.isna(row[f'answer_{lang}']):
                            setattr(existing_faq, f'answer_{lang}', row[f'answer_{lang}'])

                    existing_faq.display_order = display_order
                    existing_faq.save()
                else:
                    # Create new FAQ
                    faq_data = {
                        'tour': tour,
                        'question': row['question_en'],  # Default question
                        'answer': row['answer_en'],      # Default answer
                        'display_order': display_order
                    }

                    # Add translatable fields
                    for lang in ['en', 'ar', 'fr', 'de']:
                        if f'question_{lang}' in row and not pd.isna(row[f'question_{lang}']):
                            faq_data[f'question_{lang}'] = row[f'question_{lang}']
                        if f'answer_{lang}' in row and not pd.isna(row[f'answer_{lang}']):
                            faq_data[f'answer_{lang}'] = row[f'answer_{lang}']

                    faq = TourFAQ(**faq_data)
                    faq.save()
                    faqs_created += 1

            except Exception as e:
                errors.append(f"FAQ Row {index+1}: {str(e)}")

        return faqs_created

    # Now let's update the import_tours method to use these processing methods
    def import_tours(self, request):
        if request.method == "POST":
            form = TourImportForm(request.POST, request.FILES)
            if form.is_valid():
                file_format = form.cleaned_data['file_format']
                tour_file = request.FILES['file']
                create_missing_destinations = form.cleaned_data['create_missing_destinations']
                create_missing_categories = form.cleaned_data['create_missing_categories']
                update_existing = form.cleaned_data['update_existing']

                # Check file extension
                if file_format == 'csv' and not tour_file.name.endswith('.csv'):
                    messages.error(request, _("Wrong file format. Please upload a CSV file."))
                    return redirect('admin:tour_tour_changelist')

                if file_format == 'xlsx' and not tour_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, _("Wrong file format. Please upload an Excel file."))
                    return redirect('admin:tour_tour_changelist')

                # Process the file
                try:
                    # Start transaction to ensure all-or-nothing import
                    with transaction.atomic():
                        tours_created = 0
                        tours_updated = 0
                        itineraries_created = 0
                        faqs_created = 0
                        errors = []

                        # Read the file
                        if file_format == 'csv':
                            # CSV file - all data in one file
                            data = pd.read_csv(tour_file, encoding='utf-8')
                            # Process as a single sheet
                            tours_created, tours_updated = self._process_tours_data(
                                data,
                                create_missing_destinations,
                                create_missing_categories,
                                update_existing,
                                tours_created,
                                tours_updated,
                                errors
                            )
                        else:
                            # Excel file - multiple sheets
                            xl = pd.ExcelFile(tour_file)

                            # Process Tours sheet
                            if 'Tours' in xl.sheet_names:
                                tours_data = pd.read_excel(xl, 'Tours')
                                tours_created, tours_updated = self._process_tours_data(
                                    tours_data,
                                    create_missing_destinations,
                                    create_missing_categories,
                                    update_existing,
                                    tours_created,
                                    tours_updated,
                                    errors
                                )
                            else:
                                errors.append(_("Excel file must contain a 'Tours' sheet"))

                            # Process Itinerary sheet
                            if 'Itinerary' in xl.sheet_names:
                                itinerary_data = pd.read_excel(xl, 'Itinerary')
                                itineraries_created = self._process_itinerary_data(itinerary_data, errors)

                            # Process FAQs sheet
                            if 'FAQs' in xl.sheet_names:
                                faq_data = pd.read_excel(xl, 'FAQs')
                                faqs_created = self._process_faq_data(faq_data, errors)

                        # Show success message
                        success_messages = []
                        if tours_created > 0:
                            success_messages.append(_(f"{tours_created} tours created"))
                        if tours_updated > 0:
                            success_messages.append(_(f"{tours_updated} tours updated"))
                        if itineraries_created > 0:
                            success_messages.append(_(f"{itineraries_created} itinerary days created"))
                        if faqs_created > 0:
                            success_messages.append(_(f"{faqs_created} FAQs created"))

                        if success_messages:
                            # Convert lazy objects to strings before joining (Consistency fix)
                            messages.success(request, str(_("Import successful: ")) + ", ".join(map(str, success_messages)))
                        else:
                            messages.warning(request, _("Import completed, but no data was imported."))

                        # Show errors if any
                        if errors:
                            # Relying on top-level import for mark_safe
                            error_message = _("Errors occurred during import:") + "<br>" + "<br>".join(errors)
                            messages.error(request, mark_safe(error_message))

                except Exception as e:
                    messages.error(request, _(f"Error processing file: {str(e)}"))

                return redirect('admin:tour_tour_changelist')
        else:
            form = TourImportForm()

        context = {
            'form': form,
            'title': _("Import Tours"),
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
        }
        return render(request, 'admin/tour/tour/import_form.html', context)


@admin.register(TourGuide)
class TourGuideAdmin(TranslationAdmin):
    list_display = ('user', 'years_experience', 'rating', 'is_active')
    list_filter = ('speciality', 'is_active')
    search_fields = ('user__first_name', 'user__last_name', 'user__email', 'bio')
    filter_horizontal = ('speciality',)


@admin.register(Promotion)
class PromotionAdmin(TranslationAdmin):
    list_display = ('title', 'code', 'discount_percentage', 'start_date', 'end_date', 'is_active')
    list_filter = ('is_active', 'start_date', 'end_date')
    search_fields = ('title', 'code', 'description')
    filter_horizontal = ('tours',)
